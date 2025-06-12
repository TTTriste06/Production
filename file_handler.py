import pandas as pd
from io import BytesIO
import pandas as pd
import streamlit as st
import copy
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta


TARGET_COLUMNS = [
    "订单号", "封装厂", "封装形式", "waferin", "需排产",
    "排产周期", "磨划周期", "封装周期", "总产能", "分配产能", "实际开始测试日期"
]


def extract_order_info(uploaded_file):
    """
    从上传的 Excel 中提取 Sheet1 的目标字段。
    
    参数:
        uploaded_file: Streamlit 上传的文件对象

    返回:
        pd.DataFrame: 包含目标字段的数据
    """
    try:
        df = pd.read_excel(uploaded_file, sheet_name="Sheet1", header=3)
        matching_columns = [col for col in TARGET_COLUMNS if col in df.columns]
        df_filtered = df[matching_columns].copy()
        return df_filtered
    except Exception as e:
        return pd.DataFrame({"错误信息": [str(e)]})
        

def adjust_column_width_for_openpyxl(ws, df, start_col=25):
    for i, col in enumerate(df.columns):
        col_letter = get_column_letter(start_col + i)
        content_max_len = (
            df[col].dropna().astype(str).str.len().max()
            if not df[col].dropna().empty
            else 0
        )
        header_len = len(str(col))
        width = min(max(content_max_len, header_len) * 1.2 + 7, 50)
        ws.column_dimensions[col_letter].width = width

def write_xyz_columns(excel_file: BytesIO, df_info: pd.DataFrame) -> BytesIO:
    """
    写入 X、Y、Z 列：
    - 第3/4行写入标题
    - X列填入 '预估开始测试日期'
    - Z列填入固定值 "排产"
    - 自动调整列宽
    """
    wb = load_workbook(excel_file)
    ws = wb["Sheet1"]

    # 表头写入（行3和4）
    header_map = {
        24: ["预估开始测试日期", "预估开始测试日期"],  # X
        25: ["结束日期", "结束日期"],          # Y（保留占位）
        26: ["日期", "星期"]                  # Z
    }

    for col_idx, values in header_map.items():
        for i, val in enumerate(values):
            ws.cell(row=3 + i, column=col_idx, value=val)

    # 写入 X列（第24列）内容
    for i, value in enumerate(df_info["预估开始测试日期"], start=5):
        ws.cell(row=i, column=24, value=value)

    # 写入 Z列（第26列）内容为固定值“排产”
    for i in range(5, 5 + len(df_info)):
        ws.cell(row=i, column=26, value="排产")

    # 调整列宽
    temp_df = pd.DataFrame({
        "预估开始测试日期": df_info["预估开始测试日期"],
        "排产": ["排产"] * len(df_info)
    })
    adjust_column_width_for_openpyxl(ws, temp_df, start_col=24)

    # 保存为 BytesIO 返回
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def compute_estimated_test_date(df):
    """
    计算“预估开始测试日期”：waferin 日期 + 排产周期 + 磨划周期 + 封装周期，单位为天。

    要求 df 中存在：
    - 'waferin'：日期列
    - '排产周期'、'磨划周期'、'封装周期'：整数字段，单位为天
    """
    # 复制 DataFrame 防止原地修改
    df = df.copy()

    # 确保日期格式正确
    df["waferin"] = pd.to_datetime(df["waferin"], errors="coerce")

    # 填充空周期为 0
    for col in ["排产周期", "磨划周期", "封装周期"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    # 计算总周期天数
    df["总周期天数"] = df["排产周期"] + df["磨划周期"] + df["封装周期"]

    # 计算预估开始测试日期
    df["预估开始测试日期"] = df["waferin"] + pd.to_timedelta(df["总周期天数"], unit="D")

    # 格式化为 yyyy/mm/dd 字符串
    df["预估开始测试日期"] = df["预估开始测试日期"].dt.strftime("%Y/%m/%d")

    return df


def write_calendar_headers(file_bytes, df_info, start_col_index=28, days=14):
    """
    在 Excel 文件中写入连续日期和星期表头（从 AB列开始），同时将这些日期列加入 df_info。
    
    参数:
    - file_bytes: 原始 Excel 文件的二进制数据
    - df_info: DataFrame 对象，将添加日期列
    - start_col_index: 起始列号（默认为 28，即 AB）
    - days: 要写入的连续天数（默认14天）

    返回:
    - updated_file: 新的二进制 Excel 文件
    - df_info: 添加了日期列的新 DataFrame
    """
    wb = load_workbook(filename=BytesIO(file_bytes))
    ws = wb.active

    today = datetime.today()
    date_columns = []
    
    for i in range(days):
        date = today + timedelta(days=i)
        date_str = date.strftime("%Y/%m/%d")
        weekday = date.strftime("%A")

        col_letter = get_column_letter(start_col_index + i)
        ws[f"{col_letter}1"] = date_str
        ws[f"{col_letter}2"] = weekday
        ws[f"{col_letter}1"].alignment = Alignment(horizontal="center")
        ws[f"{col_letter}2"].alignment = Alignment(horizontal="center")

        # 将日期列初始化加入 DataFrame
        df_info[date_str] = 0
        date_columns.append(date_str)

    # 保存 Excel 回内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.read(), df_info  # 返回更新后的文件和 df_info


