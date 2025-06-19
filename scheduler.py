import pandas as pd
import streamlit as st
from datetime import timedelta
import re
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


def convert_excel_date(val):
    if pd.isnull(val):
        return pd.NaT
    try:
        if isinstance(val, str) and val.strip().isdigit():
            val = float(val)
        if isinstance(val, (int, float)):
            return pd.to_datetime("1899-12-30") + pd.to_timedelta(val, unit="D")
        return pd.to_datetime(val, errors="coerce")
    except:
        return pd.NaT

def schedule_sheet(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    if df["åˆ†é…äº§èƒ½"].isnull().any():
        raise ValueError("éƒ¨åˆ†äº§å“ç¼ºå°‘â€œåˆ†é…äº§èƒ½â€å­—æ®µï¼Œè¯·æ£€æŸ¥åŸå§‹æ•°æ®ï¼")

    df["waferin"] = pd.to_datetime(df["waferin"], errors='coerce')
    df["å®é™…å¼€å§‹æµ‹è¯•æ—¥æœŸ"] = df["å®é™…å¼€å§‹æµ‹è¯•æ—¥æœŸ"].apply(convert_excel_date)
    st.write(df["å®é™…å¼€å§‹æµ‹è¯•æ—¥æœŸ"])

    def compute_start_date(row):
        if pd.notnull(row["å®é™…å¼€å§‹æµ‹è¯•æ—¥æœŸ"]):
            return row["å®é™…å¼€å§‹æµ‹è¯•æ—¥æœŸ"]
        return row["waferin"] + timedelta(days=int(row["æ’äº§å‘¨æœŸ"]) + int(row["ç£¨åˆ’å‘¨æœŸ"]) + int(row["å°è£…å‘¨æœŸ"]))

    df["æ’äº§èµ·å§‹æ—¥"] = df.apply(compute_start_date, axis=1)

    df.sort_values("æ’äº§èµ·å§‹æ—¥", inplace=True)

    records = []
    capacity_tracker = {}

    for _, row in df.iterrows():
        order_id = row["è®¢å•å·"]
        total_qty = int(row["éœ€æ’äº§"])
        group = (row["å°è£…å‚"], row["å°è£…å½¢å¼"])
        max_daily = int(row["åˆ†é…äº§èƒ½"])

        date = row["æ’äº§èµ·å§‹æ—¥"]
        remain = total_qty
        daily_output = {}

        while remain > 0:
            key = (group[0], group[1], date)
            used = capacity_tracker.get(key, 0)
            available = max_daily - used

            if available > 0:
                assign = min(available, remain)
                daily_output[date] = assign
                remain -= assign
                capacity_tracker[key] = used + assign

            date += timedelta(days=1)

        out_row = row.to_dict()
        for d, v in daily_output.items():
            out_row[d.strftime("%Y-%m-%d")] = v

        start_date = row["æ’äº§èµ·å§‹æ—¥"].strftime("%Y-%m-%d") if pd.notnull(row["æ’äº§èµ·å§‹æ—¥"]) else ""
        end_date = max(daily_output.keys()).strftime("%Y-%m-%d") if daily_output else start_date
        estimate_start = start_date

        out_row["é¢„ä¼°å¼€å§‹æµ‹è¯•æ—¥æœŸ"] = estimate_start
        out_row["ç»“æŸæ—¥æœŸ"] = end_date
        out_row["æ’äº§èµ·å§‹æ—¥"] = start_date

        reordered = {}
        for k in out_row:
            reordered[k] = out_row[k]
            if k == "æ’äº§èµ·å§‹æ—¥":
                reordered["é¢„ä¼°å¼€å§‹æµ‹è¯•æ—¥æœŸ"] = out_row["é¢„ä¼°å¼€å§‹æµ‹è¯•æ—¥æœŸ"]
                reordered["ç»“æŸæ—¥æœŸ"] = out_row["ç»“æŸæ—¥æœŸ"]
                break
        for k in out_row:
            if k not in reordered:
                reordered[k] = out_row[k]

        records.append(reordered)

    result_df = pd.DataFrame(records)

    # è‡ªåŠ¨åˆ—å®½è°ƒæ•´é€»è¾‘å°è£…ä¸º writer åå¤„ç†
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        result_df.to_excel(writer, index=False, sheet_name="æ’äº§è®¡åˆ’")
        worksheet = writer.book["æ’äº§è®¡åˆ’"]
        for i, col in enumerate(result_df.columns, 1):
            max_length = max(result_df[col].astype(str).map(len).max(), len(str(col)))
            worksheet.column_dimensions[get_column_letter(i)].width = max_length + 10
    output.seek(0)
    st.download_button("ğŸ“¥ ä¸‹è½½æ’äº§ç»“æœï¼ˆè‡ªåŠ¨åˆ—å®½ï¼‰", data=output.getvalue(), file_name="æ’äº§è®¡åˆ’ç»“æœ.xlsx")

    return result_df
