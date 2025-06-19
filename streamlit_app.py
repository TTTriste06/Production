import streamlit as st
import pandas as pd
from io import BytesIO
from scheduler import schedule_sheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
from datetime import datetime

st.set_page_config(page_title="封装排产计划生成器", layout="wide")
st.title("📦 委外封装排产软件")

uploaded_file = st.file_uploader("上传订单 Excel 文件（包含排产字段）", type=["xlsx"])

if uploaded_file:
    # ✅ 只读取 Sheet1，字段行设定为第3行（header=1）
    df_raw = pd.read_excel(uploaded_file, sheet_name="Sheet1", header=1)
    st.success("✅ 文件上传成功！开始解析...")

    # 从第5行作为字段行，第6行开始是数据Add commentMore actions
    header_row = 1
    df_raw.columns = df_raw.iloc[header_row]
    df_data = df_raw.iloc[header_row+1:].copy()
    df_data.reset_index(drop=True, inplace=True)

    # 检查必要字段
    required_columns = ["订单号", "投单数", "封装厂", "封装形式", "waferin", "需求", "需排产", "排产周期", "磨划周期", "封装周期", "总产能", "分配产能", "实际开始测试日期"]
    missing = [col for col in required_columns if col not in df_data.columns]
    
    if missing:
        st.error(f"❌ 缺少必要字段：{missing}")
    else:
        try:
            df_scheduled = schedule_sheet(df_data)
            st.success("✅ 排产完成！")
            st.dataframe(df_scheduled.head())

            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df_scheduled.to_excel(writer, sheet_name="排产计划", index=False)
                worksheet = writer.book["排产计划"]

                # 插入第一行空白（写入前调整数据行起始行）
                worksheet.insert_rows(1)

                # 写入星期信息（第1行）
                for col_idx, col_name in enumerate(df_scheduled.columns, 1):
                    try:
                        if pd.to_datetime(col_name, errors='coerce') is not pd.NaT:
                            weekday = pd.to_datetime(col_name).strftime('%A')
                            weekday_map = {
                                'Monday': '一', 'Tuesday': '二', 'Wednesday': '三',
                                'Thursday': '四', 'Friday': '五', 'Saturday': '六', 'Sunday': '日'
                            }
                            weekday = weekday_map.get(weekday, weekday)
                            worksheet.cell(row=1, column=col_idx, value=weekday)
                    except:
                        continue

                # 设置第二行 header 为蓝底白字
                header_fill = PatternFill(fill_type="solid", fgColor="4f81bd")  # 蓝色背景
                header_font = Font(color="FFFFFF", bold=True)  # 白色字体加粗
                for col_idx, col_name in enumerate(df_scheduled.columns, 1):
                    cell = worksheet.cell(row=2, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font

                for i, col in enumerate(df_scheduled.columns, 1):
                    max_len = max(df_scheduled[col].astype(str).map(len).max(), len(str(col)))
                    worksheet.column_dimensions[get_column_letter(i)].width = max_len + 10

                # 给 A-W 区域 header 以下区域填淡蓝色背景
                from openpyxl.styles import PatternFill
                data_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
                yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
                max_row = worksheet.max_row
                for row in range(3, max_row + 1):
                    for col in range(1, 24):  # A-W 即第1列到第23列
                        cell = worksheet.cell(row=row, column=col)
                        cell.fill = data_fill

                # 仅对排产日期列（即列名是日期）标黄
                date_cols = []
                for idx, col in enumerate(df_scheduled.columns):
                    try:
                        if pd.to_datetime(col, errors='coerce') is not pd.NaT:
                            date_cols.append((idx + 1, col))  # Excel 列从1开始
                    except:
                        continue
                
                # 为排产列中有数量的格子标黄
                for row in range(3, max_row + 1):
                    for col_idx, col_name in date_cols:
                        val = worksheet.cell(row=row, column=col_idx).value
                        if isinstance(val, (int, float)) and val > 0:
                            worksheet.cell(row=row, column=col_idx).fill = yellow_fill


            output.seek(0)
            st.download_button("📥 下载排产结果", data=output.getvalue(), file_name="排产计划结果.xlsx")
        except ValueError as e:
            st.error(f"❌ 排产失败：{e}")
